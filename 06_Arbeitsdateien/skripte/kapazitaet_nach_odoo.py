#!/usr/bin/env python3
"""Schätzung der Personalkapazität, die nach dem Umstieg auf Odoo frei wird.

    python3 kapazitaet_nach_odoo.py <ziel.xlsx>

Alle Mengen und Zeiten stehen als Eingabewerte in den Blättern und sind mit
Quelle bzw. Annahme hinterlegt. Die Summen sind Formeln – wer eine Annahme
ändert, sieht das Ergebnis sofort. Zeitaufnahmen der heutigen Abläufe gibt es
nicht; die Minutenwerte sind Schätzungen und als solche gekennzeichnet.
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ACCENT = "1F4E79"
GELB = "FFF2A8"
BLAU = "DCE6F1"
GRAU = "F2F2F2"

KOPF = Font(bold=True, color="FFFFFF", size=10)
FETT = Font(bold=True, size=10)
NORM = Font(size=10)
KLEIN = Font(size=9, italic=True, color="595959")
RAHMEN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def kopfzeile(ws, zeile, werte, breiten=None):
    for i, w in enumerate(werte, start=1):
        c = ws.cell(row=zeile, column=i, value=w)
        c.font = KOPF
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = RAHMEN
    if breiten:
        for i, b in enumerate(breiten, start=1):
            ws.column_dimensions[get_column_letter(i)].width = b
    ws.row_dimensions[zeile].height = 30


def zeile(ws, r, werte, fett=False, fuellung=None, formate=None):
    for i, w in enumerate(werte, start=1):
        c = ws.cell(row=r, column=i, value=w)
        c.font = FETT if fett else NORM
        c.border = RAHMEN
        c.alignment = Alignment(vertical="top", wrap_text=(i in (2, 7)))
        if fuellung:
            c.fill = PatternFill("solid", fgColor=fuellung)
        if formate and formate.get(i):
            c.number_format = formate[i]


def titelzeile(ws, r, text, spalten=7):
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(bold=True, size=13, color=ACCENT)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=spalten)


def hinweis(ws, r, text, spalten=7):
    c = ws.cell(row=r, column=1, value=text)
    c.font = KLEIN
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=spalten)
    ws.row_dimensions[r].height = 28


# ---------------------------------------------------------------- Annahmen
ANNAHMEN = [
    ["A1", "Anlagen im Bestand des Altsystems", 938, "Stück",
     "Datenbank `bt_kunden_anlagen`, Stand 22.08.2026 – Testdatensätze noch enthalten"],
    ["A2", "Prüfturnus im Mittel", 2.0, "Jahre",
     "VDI 6022: je nach Anlagenart ein- bis dreijährig. Konservativ mittig angesetzt"],
    ["A3", "Anlagenprüfungen je Jahr", "=$C$5/$C$6", "Stück",
     "abgeleitet aus A1 und A2. Zum Vergleich: bei Volllast nennt biotec 100 Anlagen je Woche"],
    ["A4", "Gutachten je Jahr", 260, "Stück",
     "Schätzung: mehrere Anlagen je Gutachten. Mit Michael Brand zu bestätigen"],
    ["A5", "Rechnungen je Jahr", 520, "Stück",
     "Schätzung: Prüf- und Schulungsgeschäft. Aus den Ausgangsrechnungen Mai bis August belegbar"],
    ["A6", "Kurse je Jahr", 40, "Stück",
     "Schätzung: bundesweite VDI-Seminare. Aus Kurskatalog und Seminarkalender zu bestätigen"],
    ["A7", "Teilnehmer je Kurs", 12, "Personen",
     "Durchführung ab 10 Teilnehmern (Discovery Call)"],
    ["A8", "Neukunden je Jahr", 60, "Stück",
     "Schätzung aus 10.387 Interessenten und 243 Kunden. Zu bestätigen"],
    ["A9", "Monatsabschlüsse je Jahr", 12, "Stück", "Standard"],
    ["A10", "Meldungen an LucaNet je Jahr", 12, "Stück",
     "Melderhythmus der Certania-Gruppe noch offen (Frage 46)"],
    ["A11", "Vollkostensatz je Stunde", 55.0, "EUR",
     "**Annahme.** Interne Stunden- und Kostensätze sind noch nicht geliefert (Position K3). "
     "Wert bitte durch den tatsächlichen ersetzen"],
    ["A12", "Arbeitsstunden je Vollzeitstelle und Jahr", 1600, "Stunden",
     "Netto nach Urlaub, Feiertagen und Krankheit"],
]

# Tätigkeit, Mengenbezug, Menge (Formel), Min heute, Min nachher, Rolle, Begründung
TAETIGKEITEN = [
    ["Anlagenliste erstellen und ausdrucken", "je Prüfung", "A3", 25, 5,
     "Projektleitung / Labor",
     "Heute in einer Zusatzsoftware erstellt und auf Papier ausgedruckt. Künftig aus dem "
     "Anlagenstamm in Odoo erzeugt"],
    ["Ergebnisse aus dem Papierdokument zurückführen", "je Prüfung", "A3", 20, 5,
     "Labor",
     "Auszählung und Fotos werden heute manuell am Datensatz hinterlegt"],
    ["Fotos zuordnen", "je Prüfung", "A3", 12, 4,
     "Labor / Probenehmer",
     "Fotos vom Pixel 8 werden heute per Hand dem Datensatz zugeordnet; keine Anbindung im "
     "Altsystem vorhanden"],
    ["Wiederholungsprüfungen terminieren und nachverfolgen", "je Prüfung", "A3", 15, 2,
     "Projektleitung",
     "Heute außerhalb des Systems in einer Ordnerstruktur. Der Anlagenstamm kennt kein "
     "Intervall- und kein Fälligkeitsfeld"],
    ["Angebot schreiben", "je Prüfung", "A3", 18, 6,
     "Kaufmännisch",
     "Künftig aus der fälligen Wartungsanfrage heraus mit Preis aus der Preisliste"],
    ["Rechnung erstellen und versenden", "je Rechnung", "A5", 15, 4,
     "Kaufmännisch",
     "Heute manuell nach Fertigstellung des Gutachtens angestoßen"],
    ["Zahlungseingänge zuordnen und Mahnungen schreiben", "je Rechnung", "A5", 8, 2,
     "Buchhaltung", "Künftig über Bankabgleich und Mahnstufen"],
    ["Stammdaten mehrfach pflegen", "je Neukunde", "A8", 20, 6,
     "Kaufmännisch",
     "Heute getrennt in Altsystem, Excel und Buchhaltung. Künftig einmal in Odoo, das "
     "Altsystem liest über die Stammdatenbrücke"],
    ["Kursorganisation: Anmeldung, Bestätigung, Erinnerung", "je Teilnehmer",
     "A6*A7", 9, 2,
     "Kaufmännisch",
     "Heute vollständig außerhalb eines Systems – im Altsystem gibt es keine "
     "Schulungsfunktion"],
    ["Teilnehmerliste und Bescheinigungen", "je Kurs", "A6", 90, 20,
     "Kaufmännisch", "Künftig aus dem Kurs erzeugt"],
    ["Belege für die Buchhaltung aufbereiten", "je Monat", "A9", 240, 60,
     "Buchhaltung", "Künftig DATEV-Export aus Odoo"],
    ["Monatsabschluss und Auswertungen zusammenstellen", "je Monat", "A9", 300, 120,
     "Buchhaltung / GF", "Voraussichtlich heute in Excel"],
    ["Meldung an die Gruppe aufbereiten", "je Meldung", "A10", 180, 30,
     "Buchhaltung",
     "Nach Einrichtung der LucaNet-Lieferung nahe null. Option O2 im Angebot"],
    ["Auskunft geben: Wo steht der Auftrag, was ist fällig", "je Prüfung", "A3", 6, 1,
     "alle", "Heute Rückfragen an einzelne Personen, künftig im System sichtbar"],
]

ROLLEN = [
    ["Projektleitung Hygieneinspektion", 2, "Michael Brand, Stephan Krol",
     "Anlagenlisten, Terminierung, Kundenorganisation, Gutachtenübergabe"],
    ["Labor und Probenahme", 8, "10 Probennehmer mit Laboreingabe, 2 Praktikanten",
     "Probenahme vor Ort, Bebrütung, Auszählung, Fotodokumentation"],
    ["Kaufmännisch und Organisation", 1, "Nicole Krupa",
     "Rechnungen, Schulungen, Werbung"],
    ["Buchhaltung", 1, "Annette Krupa", "Belege, Abschlüsse, Meldungen"],
    ["Laborkoordination und Prozesse", 1, "Melanie Frank",
     "Laborkoordination, übergeordnete Prozesse, Hygieneinspektion"],
    ["IT", 0, "Westbomke EDV (extern)", "Betrieb, Altsystem"],
]

VERWENDUNG = [
    ["Zweites Dozententeam aufbauen", "Wachstum",
     "biotec baut ein zweites Dozententeam auf. Freie Zeit in der Kursorganisation trägt "
     "zusätzliche Kurse ohne Neueinstellung"],
    ["Wiederholungsprüfungen konsequent nachfassen", "Umsatz",
     "Der stärkste Hebel: jede nicht terminierte Wiederholungsprüfung ist ein verlorener "
     "Auftrag. 938 Anlagen im Bestand, systematische Fälligkeitsverfolgung erstmals möglich"],
    ["Akquise aus dem Interessentenbestand", "Umsatz",
     "10.387 Interessenten liegen im Altsystem und werden heute kaum systematisch bearbeitet"],
    ["Analytik Mittweida ausbauen", "Wachstum",
     "Zweiter Geschäftszweig mit 2–3 Personen; Kapazität in der Verwaltung entlastet den Standort"],
    ["Vertretbarkeit bei Personalausfall", "Risiko",
     "Nicht monetär, aber im Betrieb wertvoll: Abläufe im System statt in Köpfen"],
]


def build(ziel):
    wb = Workbook()

    # ---------------------------------------------------------- Übersicht
    ws = wb.active
    ws.title = "Übersicht"
    titelzeile(ws, 1, "Freiwerdende Personalkapazität nach dem Umstieg auf Odoo")
    hinweis(ws, 2, "biotec GmbH · Schätzung, Stand 01.09.2026 · CertoClav Sterilizer GmbH. "
                   "Alle Mengen und Zeiten sind Eingabewerte im Blatt „Annahmen" + chr(8220) +
                   " bzw. „Tätigkeiten" + chr(8220) + " und mit Quelle hinterlegt.")
    hinweis(ws, 3, "Zeitaufnahmen der heutigen Abläufe liegen nicht vor. Die Minutenwerte sind "
                   "Schätzungen und in einem Termin mit Michael Brand und Nicole Krupa zu "
                   "bestätigen. Konservativ gerechnet: Der Wert nachher ist nie null, weil "
                   "auch ein System bedient werden muss.")
    kopfzeile(ws, 5, ["Kennzahl", "Wert", "Einheit", "Hinweis", "", "", ""],
              [46, 14, 12, 60, 12, 12, 12])
    z = 6
    for txt, formel, einheit, hint in [
        ("Gebundene Stunden je Jahr heute", "=Tätigkeiten!H20", "Stunden", "Summe aus Blatt Tätigkeiten"),
        ("Gebundene Stunden je Jahr nach Odoo", "=Tätigkeiten!I20", "Stunden", ""),
        ("Freiwerdende Stunden je Jahr", "=Tätigkeiten!J20", "Stunden", "Differenz"),
        ("Entspricht Vollzeitstellen", "=J6_/Annahmen!C16", "VZÄ",
         "bei den im Blatt Annahmen genannten Jahresstunden"),
        ("Bewertet mit dem Vollkostensatz", "=J6_*Annahmen!C15", "EUR je Jahr",
         "Vollkostensatz ist eine Annahme – bitte ersetzen"),
        ("Anteil der freiwerdenden Zeit", "=Tätigkeiten!J20/Tätigkeiten!H20", "Prozent",
         "gemessen an den heute gebundenen Stunden"),
    ]:
        zeile(ws, z, [txt, formel, einheit, hint], fuellung=GELB if z == 8 else None,
              formate={2: "#,##0.0" if einheit != "EUR je Jahr" else "#,##0 €"})
        if einheit == "Prozent":
            ws.cell(row=z, column=2).number_format = "0,0 %"
        z += 1
    # Hilfszellen für die Kettenformeln
    ws["J6"] = "=Tätigkeiten!J20"
    ws["J6"].font = Font(size=8, color="FFFFFF")
    for r in range(6, 12):
        f = ws.cell(row=r, column=2).value
        if isinstance(f, str):
            ws.cell(row=r, column=2).value = f.replace("J6_", "$J$6")

    hinweis(ws, 13, "Lesart: Die Zahl ist keine Personaleinsparung. biotec baut ein zweites "
                    "Dozententeam auf – die Zeit fließt in Wachstum. Das Blatt "
                    "„Verwendung" + chr(8220) + " nennt die vorgesehenen Ziele.")
    hinweis(ws, 14, "Empfindlichkeit: Die stärksten Treiber sind die Zahl der Anlagenprüfungen "
                    "je Jahr (Annahme A3) und der Vollkostensatz (A11). Beide sind noch nicht "
                    "belegt. Wer die Zahl belastbar braucht, klärt zuerst diese zwei Werte.")

    # ---------------------------------------------------------- Annahmen
    ws = wb.create_sheet("Annahmen")
    titelzeile(ws, 1, "Annahmen und Mengen", 5)
    hinweis(ws, 2, "Jede Zeile ist ein Eingabewert. Quelle oder Annahme steht daneben. "
                   "Werte, die als Annahme gekennzeichnet sind, sollten vor der Verwendung "
                   "gegenüber der Gruppe bestätigt werden.", 5)
    kopfzeile(ws, 4, ["Nr.", "Größe", "Wert", "Einheit", "Quelle bzw. Annahme"],
              [7, 44, 14, 12, 78])
    r = 5
    for nr, name, wert, einheit, quelle in ANNAHMEN:
        zeile(ws, r, [nr, name, wert, einheit, quelle],
              fuellung=GELB if nr in ("A11",) else None,
              formate={3: "#,##0.00" if isinstance(wert, float) else "#,##0"})
        r += 1

    # ---------------------------------------------------------- Tätigkeiten
    ws = wb.create_sheet("Tätigkeiten")
    titelzeile(ws, 1, "Tätigkeiten: heute gebundene und künftig benötigte Zeit", 10)
    hinweis(ws, 2, "Rechenweg je Zeile: Menge × Minuten ÷ 60 = Stunden je Jahr. "
                   "Nur Tätigkeiten, die einem konkreten Arbeitsschritt zugeordnet sind – "
                   "kein pauschaler Effizienzgewinn.", 10)
    kopfzeile(ws, 4, ["Tätigkeit", "Mengenbezug", "Menge je Jahr", "Min heute",
                      "Min nachher", "Rolle", "Begründung", "Std heute",
                      "Std nachher", "Std frei"],
              [42, 14, 13, 10, 11, 22, 62, 11, 12, 11])
    # Annahme-Nummer -> Zelle im Blatt Annahmen (Zeile 5 ist die erste Annahme)
    zellen = {nr: f"Annahmen!$C${5 + i}" for i, (nr, *_ ) in enumerate(ANNAHMEN)}

    def mengenformel(ausdruck):
        teile = ausdruck.split("*")
        return "=" + "*".join(zellen[x] for x in teile)

    r = 5
    for name, bezug, menge, mh, mn, rolle, grund in TAETIGKEITEN:
        zeile(ws, r, [name, bezug, mengenformel(menge), mh, mn, rolle, grund,
                      f"=C{r}*D{r}/60", f"=C{r}*E{r}/60", f"=H{r}-I{r}"],
              formate={3: "#,##0", 8: "#,##0.0", 9: "#,##0.0", 10: "#,##0.0"})
        r += 1
    zeile(ws, 20, ["Summe", "", "", "", "", "", "", "=SUM(H5:H18)", "=SUM(I5:I18)",
                   "=SUM(J5:J18)"], fett=True, fuellung=GELB,
          formate={8: "#,##0.0", 9: "#,##0.0", 10: "#,##0.0"})

    # ---------------------------------------------------------- Rollen
    ws = wb.create_sheet("Rollen")
    titelzeile(ws, 1, "Verteilung auf die Rollen", 5)
    hinweis(ws, 2, "Die Stellenzahlen sind aus den 17 Benutzerkonten des Altsystems und dem "
                   "Discovery Call abgeleitet. Ein Organigramm liegt nicht vor – die Zuordnung "
                   "ist mit biotec zu bestätigen.", 5)
    kopfzeile(ws, 4, ["Rolle", "Personen (geschätzt)", "Namentlich bekannt",
                      "Aufgaben heute", "Freiwerdende Stunden je Jahr"],
              [36, 20, 40, 58, 26])
    r = 5
    for rolle, anz, namen, aufg in ROLLEN:
        zeile(ws, r, [rolle, anz, namen, aufg,
                      f'=SUMIF(Tätigkeiten!$F$5:$F$18,A{r},Tätigkeiten!$J$5:$J$18)'],
              formate={5: "#,##0.0"})
        r += 1
    zeile(ws, 12, ["Nicht eindeutig zugeordnet", "", "", "übergreifende Tätigkeiten",
                   "=Tätigkeiten!J20-SUM(E5:E10)"], fett=True, fuellung=BLAU,
          formate={5: "#,##0.0"})
    hinweis(ws, 14, "Die Zuordnung folgt der Spalte „Rolle" + chr(8220) + " im Blatt "
                    "Tätigkeiten. Wo die Bezeichnungen nicht genau übereinstimmen, "
                    "erscheint der Rest in der Zeile „nicht eindeutig zugeordnet" + chr(8220) +
                    ". Für eine belastbare Rollenrechnung sind die Bezeichnungen mit dem "
                    "Organigramm abzugleichen.", 5)

    # ---------------------------------------------------------- Verwendung
    ws = wb.create_sheet("Verwendung")
    titelzeile(ws, 1, "Wohin die freie Kapazität fließen soll", 3)
    hinweis(ws, 2, "Bewusst keine Personaleinsparung. Diese Darstellung ist im "
                   "IT-Spend-Meeting stärker und im Betrieb ehrlicher: mehr Geschäft mit "
                   "demselben Team.", 3)
    kopfzeile(ws, 4, ["Ziel", "Art des Hebels", "Begründung"], [40, 16, 92])
    r = 5
    for v_ziel, art, grund in VERWENDUNG:
        zeile(ws, r, [v_ziel, art, grund])
        r += 1

    wb.save(ziel)
    print("geschrieben:", ziel)


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else "Kapazitaet_nach_Odoo_biotec.xlsx")
